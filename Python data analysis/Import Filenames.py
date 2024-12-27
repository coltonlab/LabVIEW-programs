# Libraries
import openpyxl
import warnings
import pyautogui
import time
import os
import tkinter as tk
from tkinter import filedialog

# Suppress specific warning from pywinauto
warnings.filterwarnings("ignore", message="Revert to STA COM threading mode")

""" This is the section that saves and writes to the excel file """
def save_and_close_excel(file_name):
    """Saves and closes the specified Excel workbook."""
    try:
        from pywinauto import Application
        app = Application().connect(title_re=file_name, timeout=10)
        app_dialog = app.window(title_re=file_name)
        app_dialog.set_focus()
        
        # Use pywinauto to interact with Excel window
        app_dialog.type_keys("^s")  # Ctrl+S to save
        # time.sleep(1)  # Ensure save completes
        app_dialog.type_keys("^w")  # Ctrl+W to close
        # time.sleep(1)  # Ensure close completes
    except Exception as e:
        print(f"Error saving or closing Excel: {e}")

def open_excel(file_path):
    """Opens the specified Excel workbook."""
    try:
        os.startfile(file_path)
        # time.sleep(1)  # Wait for Excel to open
    except Exception as e:
        print(f"Error opening Excel: {e}")

def find_last_row(sheet):
    """Find the last row with data in the specified sheet."""
    for row in range(sheet.max_row, 0, -1):
        if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row
    return 1  # If no data found, start from the first row

def append_to_excel(file_path, sheet_name, entries):
    """Append values to the next available row in the specified Excel sheet."""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    
    # Find the last used row
    last_row = find_last_row(sheet)
    
    i = 1
    for values in entries:
        # Append values to the next row
        for col, value in enumerate(values, start=1):
            sheet.cell(row=last_row + i, column=col).value = value
        i = i + 1
    
    # Save the workbook
    wb.save(file_path)

def update_excel_file(values_to_append):
    file_path = r'C:/Data/All Scan Notes New.xlsx'
    sheet_name = 'Sheet1'
    file_name = 'All Scan Notes New.xlsx'  # Title of the Excel window

    """Handles the entire process of saving, closing, writing to, and reopening the Excel file."""
    try:
        # save_and_close_excel(file_name)
        append_to_excel(file_path, sheet_name, values_to_append)
        # open_excel(file_path)
    except Exception as e:
        print(f"Error updating Excel file: {e}")

""" This opens up the prompt window to ask for the folder with all of the filenames """
def select_folder():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    folder_selected = filedialog.askdirectory()  # Open folder selection dialog
    return folder_selected

def list_folder_contents(folder_path):
    entries = []
    with os.scandir(folder_path) as it:

        for entry in it:
            info = entry.stat()
            name = f"{entry.name}"
            date = f"{os.path.basename(folder_path)}"
            created = f"{time.strftime('%I:%M %p', time.localtime(info.st_birthtime))}"
            
            if name[-4:] == ".xls":            
                values = [date,created,"","","",name]
                entries.append(values)

    return entries

    with open(file_path, 'w') as f:
        for item in entries:
            f.write(f"{item['date']}\t{item['created']}\t\t\t\t{item['name']}\n")

def format_folder_contents(folder_contents):
    entries = []
    for item in folder_contents:
        entries.append(f"{item['date']}\t{item['created']}\t\t\t\t{item['name']}")

if __name__ == "__main__":
    folder_path = select_folder()
    if folder_path:
        folder_contents = list_folder_contents(folder_path)
        output_file = os.path.join(folder_path, 'folder_contents.xls')
        update_excel_file(folder_contents)
    else:
        print("No folder selected")










