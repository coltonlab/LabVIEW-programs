# Libraries
import openpyxl
import warnings
import pyautogui
import time
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from datetime import datetime

# Suppress specific warning from pywinauto
warnings.filterwarnings("ignore", message="Revert to STA COM threading mode")

def create_new_excel(file_path, sheet_name, entries):
    """Create a new Excel file and write the entries."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    # Optional: Add headers
    headers = ['Folder Name', 'Created Time', '', '', '', 'File Name']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col).value = header

    # Write entries starting from the second row (first row = headers)
    for i, values in enumerate(entries, start=2):
        for col, value in enumerate(values, start=1):
            sheet.cell(row=i, column=col).value = value

    # Save the new workbook
    wb.save(file_path)
    print(f"New Excel file created at: {file_path}")

def select_folder():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    folder_selected = filedialog.askdirectory()  # Open folder selection dialog
    return folder_selected

def list_folder_contents(folder_path):
    entries = []
    with os.scandir(folder_path) as it:
        i = 1
        for entry in it:
            info = entry.stat()
            name = f"{entry.name}"
            date = f"{os.path.basename(folder_path)}"
            created_timestamp = info.st_birthtime  # Raw timestamp for sorting
            created_formatted = f"{time.strftime('%I:%M %p', time.localtime(created_timestamp))}"
            # print(time.localtime(created_timestamp))
            i += 1  # Increment to ensure unique timestamps for sorting
            if name.lower().endswith(".xls"):
                values = [date, created_formatted, "", "", "", name]
                entries.append((created_timestamp, values))  # Store tuple (timestamp, values)

    # Sort the entries by creation time
    entries.sort(key=lambda x: x[0])

    # Extract just the sorted value lists
    sorted_entries = [entry[1] for entry in entries]
    return sorted_entries

def open_excel(file_path):
    """Opens the specified Excel workbook."""
    try:
        os.startfile(file_path)
        # time.sleep(1)  # Wait for Excel to open
    except Exception as e:
        print(f"Error opening Excel: {e}")

if __name__ == "__main__":
    folder_path = select_folder()
    if folder_path:
        folder_contents = list_folder_contents(folder_path)
        
        # Generate a new file name with timestamp
        # timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_file_name = f'folder_contents.xlsx'
        output_file = os.path.join(folder_path, new_file_name)
        
        # Create the new Excel file
        create_new_excel(output_file, 'Sheet1', folder_contents)
        open_excel(output_file)
    else:
        print("No folder selected")
